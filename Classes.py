import json
import openpyxl
from openpyxl import Workbook


class ProcessData:

    def __init__(self, input_list):
        self.input_list = input_list

    def change_list_to_dict(self):
        final_dict = {}
        for value in self.input_list:
            final_dict[value] = value
        return final_dict

    def save_json(self, final_dict):
        with open('file1.json', 'w') as json_file:
            json.dump(final_dict, json_file)
            json_file.close()

    def get_data_from_json(self):
        f = open('file1.json')
        data = json.load(f)
        return data

    def create_excel_file(self, data):
        wb = Workbook()
        ws = wb.active
        ws.title = 'New page'
        for index, key in enumerate(data.keys()):
            ws.cell(row=(index+1), column=1, value=key)
        for index, value in enumerate(data.values()):
            ws.cell(row=(index+1), column=2, value=value)
        wb.save('sample.xlsx')

    def prepare_data_for_excel(self):
        dict_data = self.change_list_to_dict()
        self.save_json(dict_data)
        dict_data_from_json = self.get_data_from_json()
        return dict_data_from_json

    def get_data_from_excel(self):
        _dict = {}
        wb_obj = openpyxl.load_workbook('sample.xlsx')
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column
        for key in sheet_obj.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            _dict[key[0]] = key[1]
        print(_dict)


process_data = ProcessData([1, 2, 3, 'dvd'])
dict_for_excel = process_data.prepare_data_for_excel()
process_data.create_excel_file(dict_for_excel)
process_data.get_data_from_excel()

